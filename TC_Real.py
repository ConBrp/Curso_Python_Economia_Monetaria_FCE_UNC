import pandas as pd
import json
import requests
import datetime
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.styles import Font, PatternFill

df = pd.read_excel("https://www.bcra.gob.ar/Pdfs/PublicacionesEstadisticas/ITCRMSerie.xlsx", header=[1],
				   sheet_name=["ITCRM y bilaterales", "ITCRM y bilaterales prom. mens."])

TCRD = df["ITCRM y bilaterales"][["Período", "ITCRM "]].dropna().copy()
TCRM = df["ITCRM y bilaterales prom. mens."][["Período", "ITCRM "]].dropna().copy()

TCRD["Período"] = pd.to_datetime(TCRD["Período"])
TCRD["Dia"] = TCRD["Período"].dt.day
TCRD["Mes"] = TCRD["Período"].dt.month
TCRD["Año"] = TCRD["Período"].dt.year
TCRD["Date"] = TCRD["Mes"].astype(str) + '-' + TCRD["Año"].astype(str)
TCRD["Date"] = pd.to_datetime(TCRD['Date'], format='%m-%Y').dt.strftime("%m-%Y")

TCRM["Período"] = pd.to_datetime(TCRM["Período"])
TCRM["Dia"] = TCRM["Período"].dt.day
TCRM["Mes"] = TCRM["Período"].dt.month
TCRM["Año"] = TCRM["Período"].dt.year
TCRM["Date"] = TCRM["Mes"].astype(str) + '-' + TCRM["Año"].astype(str)
TCRM["Date"] = pd.to_datetime(TCRM['Date'], format='%m-%Y').dt.strftime("%m-%Y")

# IPC = pd.read_excel("https://www.ieric.org.ar/wp-content/uploads/2024/01/IPC-Prov-San-Luis.xlsx", header=3)[["Periodo","Nivel General"]].dropna().copy()
IPC = pd.read_excel("https://www.ieric.org.ar/wp-content/uploads/2020/01/IPC-Prov-San-Luis.xlsx", header=3)[["Periodo","Nivel General"]].dropna().copy()

IPC["Periodo"] = pd.to_datetime(IPC["Periodo"])
IPC["Dia"] = IPC["Periodo"].dt.day
IPC["Mes"] = IPC["Periodo"].dt.month
IPC["Año"] = IPC["Periodo"].dt.year
IPC["Date"] = IPC["Mes"].astype(str) + '-' + IPC["Año"].astype(str)
IPC["Date"] = pd.to_datetime(IPC['Date'], format='%m-%Y').dt.strftime("%m-%Y")
IPC = IPC[["Nivel General", "Date"]].copy()
IPC.columns = ["IPC", "Date"]

headers = {'Content-type': 'application/json'}
inicio = 2005
final = datetime.date.today().year
CPIUS = pd.DataFrame(columns=["Año", "CPI", "Mes", "Date"])
while inicio <= final:
	data = json.dumps({"seriesid": ['CUUR0000SA0'], "startyear": str(inicio), "endyear": str(inicio + 9)})
	response = requests.post('https://api.bls.gov/publicAPI/v2/timeseries/data/', data=data, headers=headers)
	json_data = json.loads(response.text)
	if not json_data["Results"]["series"][0]["data"]:
		break
	CPI = pd.DataFrame(json_data["Results"]["series"][0]["data"]).iloc[:, :-1]
	CPI["Mes"] = CPI["period"].apply(lambda x: x[1:])
	CPI = CPI[["year", "value", "Mes"]].copy()
	CPI.columns = ["Año", "CPI", "Mes"]
	CPI["Date"] = CPI["Mes"].astype(str) + '-' + CPI["Año"].astype(str)
	CPI["Date"] = pd.to_datetime(CPI['Date'], format='%m-%Y').dt.strftime("%m-%Y")
	CPIUS = pd.concat([CPI, CPIUS], ignore_index=True)
	inicio = int(CPI.iloc[0]["Año"])
	inicio += 1

CPIUS = CPIUS[["CPI", "Date"]].copy()
CPIUS["CPI"] = pd.to_numeric(CPIUS["CPI"])

infla = pd.merge(IPC, CPIUS, how="inner", on="Date")
infla["InflaMensualAr"] = infla["IPC"].pct_change()
infla["InflaMensualUS"] = infla["CPI"].pct_change()

infla = infla[["Date", "IPC", "InflaMensualAr", "CPI", "InflaMensualUS"]]
infla.columns = ["Mes IPC/CPI", "IPC", "InflaMensualAr", "CPI", "InflaMensualUS"]
infla["Mes IPC/CPI"] = pd.to_datetime(infla["Mes IPC/CPI"])

infla.to_excel("TCR.xlsx", sheet_name="Datos", index=False, startcol=5)

df = pd.read_excel("https://www.bcra.gob.ar/Pdfs/PublicacionesEstadisticas/series.xlsm",
				   header=[0, 1, 2, 3, 4, 5, 6, 7, 8], sheet_name="RESERVAS")
df.columns = [str(i) for i in range(1, len(df.columns) + 1)]

TCN = df[df["17"] == "D"][["1", "16"]].copy()

TCN.columns = ["Período", "TCN"]

TCRD = TCRD[["Período", "ITCRM "]].copy()

TCRN = pd.merge(TCN, TCRD, how="inner", on="Período")

TCRN.columns = ["Fecha", "TCN", "ITCRM"]
TCRD.columns = ["Fecha ITCRM", "ITCRM"]
TCRM = TCRM[["Date", "ITCRM "]].copy()
TCRM.columns = ["Mes ITCRM", "ITCRM"]

with pd.ExcelWriter("TCR.xlsx", mode='a', engine='openpyxl', if_sheet_exists="overlay") as writer:
	TCRN.to_excel(writer, startcol=0, sheet_name="Datos", index=False, header=True)
	TCRM.to_excel(writer, startcol=3, sheet_name="Datos", index=False, header=True)

wb = load_workbook('TCR.xlsx')
sheet = wb["Datos"]

for celda in sheet['A']:
	celda.number_format = 'd/m/yyyy'
for celda in sheet['F']:
	celda.number_format = 'm/yyyy'

sheet = wb.create_sheet("Cálculos")

img = Image('formula.png')

img.width = 600 * .75
img.height = 200 * .75

sheet.add_image(img, 'A1')

wb.save('TCR.xlsx')

wb = load_workbook('TCR.xlsx')
sheet = wb["Datos"]
grafico = LineChart()
grafico.title = "Tipo de Cambio Real Multilateral"
grafico.x_axis.title = "Fecha"
grafico.y_axis.title = "Índice"

datos_referencia = Reference(sheet,
							 min_col=3,
							 min_row=2,
							 max_row=len(sheet["C"]))
serie = Series(datos_referencia, title="ITCRM")
categorias_referencia = Reference(sheet,
								  min_col=1,
								  min_row=2,
								  max_row=len(sheet["C"]))
grafico.height = 15
grafico.width = 30
grafico.series.append(serie)
grafico.set_categories(categorias_referencia)
sheet.add_chart(grafico, "L1")
for col in sheet.columns:
	sheet.column_dimensions[col[0].column_letter].auto_size = True
wb.save('TCR.xlsx')

wb = load_workbook('TCR.xlsx')
sheet = wb["Cálculos"]

sheet["A9"] = "Por Fecha"
sheet["A9"].font = Font(size=14, bold=True)
sheet["A10"] = "Fecha Deseada"
sheet["A11"] = "TCN"
sheet["A12"] = "TCR"

sheet["A14"] = "TCR Último"

sheet["A17"] = "CPI Inicial"
sheet["A18"] = "IPC Inicial"
sheet["A19"] = "CPI Final"
sheet["A20"] = "IPC Final"

sheet["A22"] = "Por Valor"
sheet["A22"].font = Font(size=14, bold=True)
sheet["A23"] = "TCN"
sheet["A24"] = "TCR"

sheet["A26"] = "TCR Último"

sheet["A29"] = "CPI Inicial"
sheet["A30"] = "IPC Inicial"
sheet["A31"] = "CPI Final"
sheet["A32"] = "IPC Final"

sheet["B10"].fill = PatternFill(start_color="0099CCFF", end_color="0099CCFF", fill_type="solid")
sheet["B11"] = "=VLOOKUP($B$10, Datos!A:C, 2)"
sheet["B12"] = "=VLOOKUP($B$10, Datos!A:C, 3)"

sheet["B14"] = "=VLOOKUP(MAX(Datos!C:C)+1, Datos!C:C, 1)"

sheet["B17"] = "=VLOOKUP($D$11, Datos!F:J, 4)"
sheet["B18"] = "=VLOOKUP($D$11, Datos!F:J, 2)"
sheet["B19"].fill = PatternFill(start_color="0099CCFF", end_color="0099CCFF", fill_type="solid")
sheet["B20"].fill = PatternFill(start_color="0099CCFF", end_color="0099CCFF", fill_type="solid")

sheet["B23"].fill = PatternFill(start_color="0099CCFF", end_color="0099CCFF", fill_type="solid")
sheet["B24"].fill = PatternFill(start_color="0099CCFF", end_color="0099CCFF", fill_type="solid")

sheet["B26"] = "=VLOOKUP(MAX(Datos!C:C)+1, Datos!C:C, 1)"

sheet["B29"].fill = PatternFill(start_color="0099CCFF", end_color="0099CCFF", fill_type="solid")
sheet["B30"].fill = PatternFill(start_color="0099CCFF", end_color="0099CCFF", fill_type="solid")
sheet["B31"].fill = PatternFill(start_color="0099CCFF", end_color="0099CCFF", fill_type="solid")
sheet["B32"].fill = PatternFill(start_color="0099CCFF", end_color="0099CCFF", fill_type="solid")

sheet["D10"] = "Mes IPC/CPI"
sheet["D11"] = "=DATE(YEAR(B10), MONTH(B10), 1)"
sheet["D11"].number_format = 'm/yyyy'

sheet["E14"] = "Variación TCN="
sheet["E15"] = "=(F14-G14+H14)"

sheet["E17"] = "TCN a futuro"

sheet["E26"] = "Variación TCN="
sheet["E27"] = "=(F26-G26+H26)"

sheet["E29"] = "TCN a futuro"

sheet["F14"] = "=B12/B14-1"
sheet["F17"] = "=B11*(1+E15)"

sheet["F26"] = "=B24/B26-1"
sheet["F29"] = "=B23*(1+E27)"

sheet["G14"] = "=B19/B17-1"
sheet["G26"] = "=B31/B29-1"

sheet["H1"] = "University of Michigan (Expected Change in Prices)"
sheet["H1"].hyperlink = "http://www.sca.isr.umich.edu/tables.html"
sheet["H1"].font = Font(color="000000FF", italic=True)
sheet["H2"] = "OECD (Inflation forecast)"
sheet["H2"].hyperlink = "https://data.oecd.org/price/inflation-forecast.htm"
sheet["H2"].font = Font(color="000000FF", italic=True)
sheet["H3"] = "BCRA (REM, Precios minoristas)"
sheet["H3"].hyperlink = "https://www.bcra.gob.ar/PublicacionesEstadisticas/REM-precios-minoristas.asp"
sheet["H3"].font = Font(color="000000FF", italic=True)

sheet["H14"] = "=B20/B18-1"
sheet["H26"] = "=B32/B30-1"

sheet["F9"] = "Completar las celdas de este color."
sheet["F9"].fill = PatternFill(start_color="0099CCFF", end_color="0099CCFF", fill_type="solid")
sheet.merge_cells("F9:H9")

for col in sheet.columns:
	sheet.column_dimensions[col[0].column_letter].auto_size = True

wb.save('TCR.xlsx')
